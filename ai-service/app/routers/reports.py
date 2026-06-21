import io
import csv
import json
from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

router = APIRouter()


# ─── PDF ──────────────────────────────────────────────────────────────────────

@router.post("/generate")
def generate_report(body: dict):
    kpis = body.get("kpis", {})
    insights = body.get("insights", {})
    title = body.get("title", "Business Intelligence Report")

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=inch * 0.75, leftMargin=inch * 0.75,
                            topMargin=inch, bottomMargin=inch)
    doc.build(_build_story(kpis, insights, title))
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{_safe(title)}.pdf"'})


# ─── CSV ──────────────────────────────────────────────────────────────────────

@router.post("/export/csv")
def export_csv(body: dict):
    data: list = body.get("data", [])
    filename: str = body.get("filename", "export")

    if not data:
        return StreamingResponse(iter([""]), media_type="text/csv")

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=data[0].keys())
    writer.writeheader()
    writer.writerows(data)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{_safe(filename)}.csv"'}
    )


# ─── Excel ────────────────────────────────────────────────────────────────────

@router.post("/export/excel")
def export_excel(body: dict):
    data: list = body.get("data", [])
    filename: str = body.get("filename", "export")
    sheet_name: str = body.get("sheet_name", "Data")
    kpis: dict = body.get("kpis", {})

    buf = io.BytesIO()
    wb = openpyxl.Workbook()

    # ── KPI sheet (if provided) ──
    if kpis:
        ws_kpi = wb.active
        ws_kpi.title = "KPI Summary"
        header_fill = PatternFill("solid", fgColor="2563EB")
        header_font = Font(bold=True, color="FFFFFF")
        ws_kpi.append(["Metric", "Value"])
        for cell in ws_kpi[1]:
            cell.fill = header_fill
            cell.font = header_font
        kpi_labels = {
            "totalRevenue": "Total Revenue", "totalOrders": "Total Orders",
            "avgOrderValue": "Avg Order Value", "grossProfit": "Gross Profit",
            "revenueGrowthPct": "Revenue Growth %", "totalCustomers": "Total Customers",
        }
        for key, label in kpi_labels.items():
            if key in kpis:
                ws_kpi.append([label, kpis[key]])
        ws_kpi.column_dimensions['A'].width = 25
        ws_kpi.column_dimensions['B'].width = 18

    # ── Data sheet ──
    if data:
        ws_data = wb.create_sheet(title=sheet_name)
        headers = list(data[0].keys())
        ws_data.append(headers)
        header_fill2 = PatternFill("solid", fgColor="1E3A5F")
        header_font2 = Font(bold=True, color="FFFFFF")
        for cell in ws_data[1]:
            cell.fill = header_fill2
            cell.font = header_font2
            cell.alignment = Alignment(horizontal="center")
        for row in data:
            ws_data.append([row.get(h, "") for h in headers])
        for col in ws_data.columns:
            ws_data.column_dimensions[col[0].column_letter].width = 18

    if not kpis and not data:
        wb.active.title = "Empty"

    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{_safe(filename)}.xlsx"'}
    )


def _safe(name: str) -> str:
    return name.replace(" ", "_").replace("/", "-")[:60]


def _build_story(kpis: dict, insights: dict, title: str) -> list:
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=18, textColor=colors.HexColor("#1e3a5f"))
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=14, textColor=colors.HexColor("#2563eb"))
    body_style = styles["BodyText"]
    story = []

    story.append(Paragraph(title, h1))
    story.append(Spacer(1, 12))

    if kpis:
        story.extend(_build_kpi_table(kpis, h2))

    if insights:
        story.extend(_build_insights_section(insights, h2, body_style))

    return story


def _build_kpi_table(kpis: dict, h2) -> list:
    styles = getSampleStyleSheet()
    story = [Paragraph("Executive KPI Summary", h2), Spacer(1, 8)]
    kpi_labels = {
        "totalRevenue": "Total Revenue (₹)",
        "totalOrders": "Total Orders",
        "avgOrderValue": "Avg. Order Value (₹)",
        "grossProfit": "Gross Profit (₹)",
        "revenueGrowthPct": "Revenue Growth (%)",
        "totalCustomers": "Total Customers"
    }
    kpi_data = [["Metric", "Value"]]
    for key, label in kpi_labels.items():
        if key in kpis:
            val = kpis[key]
            formatted = f"{val:,.2f}" if isinstance(val, float) else str(val)
            kpi_data.append([label, formatted])

    table = Table(kpi_data, colWidths=[3.5 * inch, 2.5 * inch])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(table)
    story.append(Spacer(1, 16))
    return story


def _build_insights_section(insights: dict, h2, body_style) -> list:
    story = [Paragraph("Business Insights", h2)]
    if insights.get("summary"):
        story.append(Paragraph(insights["summary"], body_style))
        story.append(Spacer(1, 8))
    sections = [
        ("Key Findings", "keyFindings"),
        ("Growth Drivers", "growthDrivers"),
        ("Risks", "risks"),
        ("Recommendations", "recommendations"),
    ]
    for section_title, key in sections:
        items = insights.get(key, [])
        if items:
            story.append(Paragraph(section_title, ParagraphStyle("s", parent=h2, fontSize=11)))
            for item in items:
                story.append(Paragraph(f"• {item}", body_style))
            story.append(Spacer(1, 6))
    return story
